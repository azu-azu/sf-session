"""Write probe results to an Excel file.

ジョブ実行結果（JobResult リスト）を 1 行 1 ジョブの Excel ファイルに書き出す。
ヘッダは黒背景・白文字、URL セルはハイパーリンク付き。
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..config import ARCHIVE_RESULT_DIR
from ..utils import format_duration
from .job_result import JobResult

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FILL = PatternFill(fill_type="solid", fgColor="000000")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# (header, width)
_COLUMNS: list[tuple[str, int]] = [
    ("No", 6),
    ("ID", 20),
    ("report name", 40),
    ("url", 52),
    ("duration", 10),
    ("status", 10),
    ("列数", 10),
    ("備考", 80),
]


def _status_label(status: str) -> str:
    if status == "failed":
        return "失敗"
    if status == "probed":
        return "確認のみ"
    return status


def _build_data_row(result: JobResult, idx: int) -> list[object]:
    """JobResult 1件分のデータ行を構築する。"""
    report_id = result.report_id or result.job_id
    report_url = result.report_url or "-"
    duration = format_duration(result.duration_seconds)
    status = _status_label(result.status)

    column_count = result.column_count
    if column_count is None and result.discovery_columns:
        column_count = len(result.discovery_columns)

    error = result.error or "-"

    return [
        idx,
        report_id,
        result.report_name or "-",
        report_url,
        duration,
        status,
        column_count if column_count is not None else "-",
        error,
    ]


def _style_sheet(ws: Worksheet, data_end_row: int, col_count: int) -> None:
    # 1行目（実行日時）を太字にする
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 2行目（ヘッダ）を黒背景・白文字・太字にする
    for cell in ws[2]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    last_col = get_column_letter(col_count)
    ws.auto_filter.ref = f"A2:{last_col}{data_end_row}"
    ws.freeze_panes = "A3"

    # 列幅を設定
    for i, (_, width) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def write_result_excel(
    results: list[JobResult],
    output_path: Path | None = None,
    run_ts: str | None = None,
) -> Path:
    """JobResult リストを Excel ファイルに書き出す。

    Output: outputs/archive/result/probe_result_{ts}.xlsx
    """
    if output_path is None:
        ts = run_ts or datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = ARCHIVE_RESULT_DIR / f"probe_result_{ts}.xlsx"

    # タイムスタンプ表示用変換
    if run_ts:
        dt = datetime.strptime(run_ts, "%Y%m%d_%H%M%S")
    else:
        dt = datetime.now()
    ts_display = dt.strftime("%Y/%m/%d %H時%M分%S秒")

    wb = Workbook()
    ws = wb.active
    ws.title = "probe_result"

    # 1行目: 実行日時
    elapsed_secs = sum(
        s.duration_seconds for s in results if s.duration_seconds is not None
    )
    elapsed = format_duration(elapsed_secs) if elapsed_secs else None
    elapsed_part = f"（elapsed: {elapsed}）" if elapsed else ""
    ws.append([f"実行開始日時: {ts_display}{elapsed_part}"])

    # 2行目: ヘッダ
    headers = [h for h, _ in _COLUMNS]
    ws.append(headers)

    # データ行
    url_col = headers.index("url") + 1
    for idx, result in enumerate(results, start=1):
        row = _build_data_row(result, idx)
        ws.append(row)
        data_row = idx + 2
        url = result.report_url
        if url and url.startswith("http"):
            cell = ws.cell(row=data_row, column=url_col)
            cell.hyperlink = url
            cell.style = "Hyperlink"

    # スタイル適用
    data_end_row = ws.max_row
    _style_sheet(ws, data_end_row, len(_COLUMNS))

    # 列名シート
    _write_columns_sheet(wb, results)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path.resolve()


def _write_columns_sheet(wb: Workbook, results: list[JobResult]) -> None:
    """列名一覧シートを追加する。1レポート1列、横方向に並べる。"""
    probed = [r for r in results if r.discovery_columns]
    if not probed:
        return

    ws = wb.create_sheet(title="columns")

    # Row 1: レポート名（ヘッダ）
    for col_idx, r in enumerate(probed, start=1):
        cell = ws.cell(row=1, column=col_idx, value=r.report_name or r.report_id)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Row 2+: 列名を縦に並べる
    for col_idx, r in enumerate(probed, start=1):
        for row_idx, col_name in enumerate(r.discovery_columns, start=2):
            ws.cell(row=row_idx, column=col_idx, value=col_name)

    # 列幅を自動調整
    for col_idx, r in enumerate(probed, start=1):
        max_len = max(
            len(str(r.report_name or r.report_id)),
            *(len(c) for c in r.discovery_columns),
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    ws.freeze_panes = "A2"
