"""マクロ格納フォルダの xlsm からジョブ定義を読み取る。

Usage:
    python -m sf_session.macro_book_reader archive
"""

from __future__ import annotations

import argparse
import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .config import PIPELINES, PipelineConfig, VALID_PIPELINES, USER_HOME, USE_HOME_FALLBACK
from .id_files import find_latest_success_ids, read_ids_file
from .utils import strip_trailing_date
from .utils_excel import SheetNotFoundError, get_cell_value, resolve_sheet

logger = logging.getLogger(__name__)

# --- 列定数（ローマ字） ---
_COL_NO = "AA"
_COL_URL = "AB"
_COL_NEW_FILENAME = "AC"
_COL_SRC_FOLDER_NAME = "AD"
_COL_ENCODE = "AE"
_COL_SKIP = "AG"

_SHEET_NAME = "SalesForce"

# データ開始行（100 = ヘッダ、101〜 = データ）
_DATA_START_ROW = 101


@dataclass
class JobEntry:
    """1行分のジョブ定義。"""

    no: str
    report_id: str | None
    has_filename: bool
    new_filename: str
    src_folder_name: str
    encode: str
    skip: str


def _extract_id(url: str | None) -> str | None:
    """URL から末尾の ID を抽出する。"""
    if not url or not isinstance(url, str):
        return None
    url = url.strip()
    if "/" in url:
        return url.rsplit("/", 1)[-1]
    return url


def _has_filename(value) -> bool:
    """new_filename の指定があるか判定する。"""
    if value is None:
        return False
    s = str(value).strip()
    return bool(s) and s.isprintable()


def _find_xlsm(directory: Path) -> Path | None:
    """ディレクトリ内の .xlsm ファイルを1つ返す。複数あれば RuntimeError。"""
    files = sorted(directory.glob("*.xlsm"))
    if not files:
        return None
    if len(files) > 1:
        names = "\n".join(f"  - {f.name}" for f in files)
        raise RuntimeError(
            f".xlsm が複数あります。使用するファイルを1つにしてください:\n{names}"
        )
    return files[0]


def read_jobs_from_xlsm(xlsm_path: Path) -> list[JobEntry]:
    """指定した xlsm パスからジョブ定義を読み取る。"""
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    try:
        ws = resolve_sheet(wb, _SHEET_NAME)

        entries: list[JobEntry] = []
        for row in range(_DATA_START_ROW, ws.max_row + 1):
            no = get_cell_value(ws, row, _COL_NO)
            url = get_cell_value(ws, row, _COL_URL)

            # No も URL も空なら終端
            if no is None and url is None:
                break

            raw_filename = get_cell_value(ws, row, _COL_NEW_FILENAME)
            has_fn = _has_filename(raw_filename)

            entries.append(JobEntry(
                no=str(no) if no is not None else "",
                report_id=_extract_id(url),
                has_filename=has_fn,
                new_filename=strip_trailing_date(str(raw_filename).strip()) if has_fn else "",
                src_folder_name=str(get_cell_value(ws, row, _COL_SRC_FOLDER_NAME) or ""),
                encode=str(get_cell_value(ws, row, _COL_ENCODE) or ""),
                skip=str(get_cell_value(ws, row, _COL_SKIP) or ""),
            ))

        return entries
    finally:
        wb.close()


def read_jobs(macro_dir: Path) -> list[JobEntry]:
    """マクロ格納フォルダの xlsm からジョブ定義を読み取る。"""
    if not macro_dir.is_dir():
        raise FileNotFoundError(f"'{macro_dir}' が見つかりません。")

    xlsm_path = _find_xlsm(macro_dir)
    if xlsm_path is None:
        raise FileNotFoundError(
            f"'{macro_dir.name}/' に .xlsm がありません。"
        )

    logger.info("xlsm から読み取り: %s", xlsm_path)
    return read_jobs_from_xlsm(xlsm_path)


def load_active_jobs(
    pipeline: PipelineConfig,
    *,
    ids_file: bool = False,
    exclude_success: bool = False,
) -> list[JobEntry]:
    """ジョブ定義を読み込み、skip / ids-file / success 除外でフィルタして返す。"""
    jobs = read_jobs(pipeline.macro_dir)
    logger.info("ジョブ定義: %d 件読み取り", len(jobs))

    active = [j for j in jobs if not j.skip]
    logger.info(
        "実行対象: %d 件 (skip 除外: %d 件)",
        len(active), len(jobs) - len(active),
    )

    if ids_file:
        target_ids = read_ids_file(pipeline.ids_file)
        if not target_ids:
            logger.warning("ids-file に ID の記載が 0 件です — %s", pipeline.ids_file)
            return []
        before = len(active)
        active = [j for j in active if j.report_id in target_ids]
        logger.info(
            "ids-file フィルタ: %d 件 → %d 件 (ids-file: %d IDs)",
            before, len(active), len(target_ids),
        )

    if exclude_success:
        ids_path = find_latest_success_ids(pipeline.result_dir)
        if ids_path is None:
            logger.warning("success_ids ファイルが見つかりません。除外なしで続行。")
        else:
            success_ids = read_ids_file(ids_path)
            before = len(active)
            active = [j for j in active if j.report_id not in success_ids]
            logger.info(
                "success 除外: %d 件 → %d 件 (%s: %d IDs)",
                before, len(active), ids_path.name, len(success_ids),
            )

    return active


def write_ids_all(jobs: list[JobEntry], ids_dir: Path) -> Path:
    """マクロファイル内の全 report_id を ids_all.txt に出力する。"""
    all_ids = [j.report_id for j in jobs if j.report_id]
    ids_dir.mkdir(parents=True, exist_ok=True)
    out = ids_dir / "ids_all.txt"
    out.write_text("\n".join(all_ids) + "\n", encoding="utf-8")
    logger.info("ids_all.txt 出力: %d 件 → %s", len(all_ids), out)
    return out


def _log_jobs(entries: list[JobEntry]) -> None:
    """ジョブ定義をテーブル形式でターミナルに表示する。"""
    header = (
        f"{'No':<6} {'ID':<20} {'filename?':<10} "
        f"{'new_filename':<30} {'src_folder_name':<30} "
        f"{'encode':<10} {'skip':<6}"
    )
    lines = [header, "-" * len(header)]

    for i, e in enumerate(entries, start=1):
        id_str = e.report_id or "(なし)"
        fn_flag = "True" if e.has_filename else "False"
        lines.append(
            f"{i:<6} {id_str:<20} {fn_flag:<10} "
            f"{e.new_filename:<30} {e.src_folder_name:<30} "
            f"{e.encode:<10} {e.skip:<6}"
        )

    lines.append(f"合計: {len(entries)} 件")
    logger.info("\n%s", "\n".join(lines))


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="マクロ格納フォルダの xlsm からジョブ定義を読み取る",
    )
    parser.add_argument(
        "pipeline",
        choices=VALID_PIPELINES,
        help="実行対象の pipeline 名",
    )
    return parser.parse_args(argv)


if __name__ == "__main__":
    from datetime import datetime as _dt

    from .utils import setup_logging

    setup_logging()
    _args = parse_args()
    _pipeline = PIPELINES[_args.pipeline]

    xlsm_path = _find_xlsm(_pipeline.macro_dir)
    if xlsm_path is None:
        logger.error("'%s/' に .xlsm がありません。", _pipeline.macro_dir)
    else:
        mtime = _dt.fromtimestamp(xlsm_path.stat().st_mtime)
        logger.info("Use Home: %s", USE_HOME_FALLBACK)
        logger.info("Home: %s", USER_HOME)
        logger.info("ファイル: %s (更新: %s)", xlsm_path, mtime.strftime("%Y-%m-%d %H:%M"))
        jobs = read_jobs_from_xlsm(xlsm_path)
        _log_jobs(jobs)

        write_ids_all(jobs, _pipeline.ids_dir)
