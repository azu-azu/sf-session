"""excel関連の共通ユーティリティ。"""

from __future__ import annotations

from difflib import get_close_matches

from openpyxl.worksheet.worksheet import Worksheet


class SheetNotFoundError(Exception):
    """xlsm に期待するシートが存在しない場合のエラー。"""


def col_to_index(col_letter: str) -> int:
    """Excel 列文字を 1-based インデックスに変換する。"""
    result = 0
    for ch in col_letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def get_cell_value(ws, row: int, col_letter: str):
    """ワークシートからセル値を取得する。"""
    return ws.cell(row=row, column=col_to_index(col_letter)).value


def resolve_sheet(wb, expected: str) -> Worksheet:
    """expected 名のシートを返す。見つからなければ typo 候補を提示して error。"""
    if expected in wb.sheetnames:
        return wb[expected]

    candidates = get_close_matches(expected, wb.sheetnames, n=3, cutoff=0.5)
    sheets_list = ", ".join(wb.sheetnames)

    if candidates:
        suggestion = ", ".join(candidates)
        raise SheetNotFoundError(
            f"シート '{expected}' が見つかりません。"
            f"\n  typo の可能性: {suggestion}"
            f"\n  → Excel でシート名を '{expected}' に修正してください。"
            f"\n  (全シート: {sheets_list})"
        )
    raise SheetNotFoundError(
        f"シート '{expected}' が見つかりません。"
        f"\n  全シート: {sheets_list}"
        f"\n  → Excel に '{expected}' シートを作成してください。"
    )
