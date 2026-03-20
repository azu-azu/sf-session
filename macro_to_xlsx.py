"""xlsm から必要なジョブ定義だけを抽出した xlsx を生成する。

Usage:
    python sf-session/macro_to_xlsx.py [--macro-dir PATH]
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .config import JOBS_XLSX, MACRO_DIR, SF_BASE_URL, create_sf_client
from .macro_book_reader import JobEntry, read_jobs_from_xlsm

logger = logging.getLogger(__name__)

_HEADERS = [
    "no",
    "report_id",
    "report_name",
    "new_filename",
    "dst_folder_name",
    "encode",
    "skip",
]


def _find_single_xlsm(macro_dir: Path) -> Path:
    """xlsm を1つだけ返す。0 or 2+ はエラー。"""
    files = list(macro_dir.glob("*.xlsm"))
    if not files:
        raise FileNotFoundError(f"'{macro_dir.name}/' に .xlsm がない。")
    if len(files) > 1:
        names = [f.name for f in files]
        raise RuntimeError(f".xlsm が複数ある: {names}。1ファイルのみ対応。")
    return files[0]


_RE_TRAILING_DATE = re.compile(r"_(\d{8})$")


def _strip_trailing_date(name: str) -> str:
    """末尾の _YYYYMMDD を除去する。日付として invalid なら何もしない。"""
    m = _RE_TRAILING_DATE.search(name)
    if not m:
        return name
    try:
        dt = datetime.strptime(m.group(1), "%Y%m%d")
    except ValueError:
        return name
    if dt.year != datetime.now().year:
        return name
    return name[: m.start()]


def _entry_to_row(no: int, entry: JobEntry, report_name: str) -> list:
    return [
        no,
        entry.report_id or "",
        report_name,
        _strip_trailing_date(entry.new_filename),
        entry.src_folder_name,
        entry.encode,
        entry.skip,
    ]


def _fetch_report_names(entries: list[JobEntry]) -> dict[str, str]:
    """SF API describe で report_name を取得。失敗時は空文字で fallback。"""
    try:
        from simple_salesforce.exceptions import SalesforceError
    except ImportError:
        logger.warning("simple-salesforce 未インストール — report_name は空欄になる")
        return {}

    try:
        sf = create_sf_client()
    except (SalesforceError, KeyError, OSError):
        logger.warning("SF 接続失敗 — report_name は空欄になる")
        return {}

    names: dict[str, str] = {}
    for entry in entries:
        rid = entry.report_id
        if not rid or rid in names:
            continue
        try:
            desc = sf.restful(f"analytics/reports/{rid}/describe", method="GET")
            meta = desc.get("reportMetadata", {})
            names[rid] = meta.get("name", "")
        except SalesforceError:
            logger.warning("describe 失敗: %s", rid)
            names[rid] = ""
    return names


def generate(macro_dir: Path = MACRO_DIR) -> Path:
    """xlsx を生成して出力パスを返す。"""
    xlsm_path = _find_single_xlsm(macro_dir)
    entries = [e for e in read_jobs_from_xlsm(xlsm_path) if e.report_id]
    now = datetime.now()

    names = _fetch_report_names(entries)

    wb = Workbook()
    ws = wb.active
    ws.title = "jobs"

    # Row 1: メタ情報
    ws["B1"] = f"出力日時: {now:%Y-%m-%d %H:%M:%S}"
    ws["E1"] = f"参照元: {xlsm_path.name}"

    # Row 2: ヘッダ
    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.cell(row=2, column=col_idx, value=header)

    # Row 3+: データ
    col_report_id = _HEADERS.index("report_id") + 1
    for no, (row_idx, entry) in enumerate(enumerate(entries, start=3), start=1):
        report_name = names.get(entry.report_id or "", "")
        for col_idx, value in enumerate(
            _entry_to_row(no, entry, report_name), start=1
        ):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == col_report_id and value:
                cell.hyperlink = f"{SF_BASE_URL}/{value}"
                cell.style = "Hyperlink"

    for col_idx in range(1, len(_HEADERS) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    output_path = macro_dir / JOBS_XLSX
    wb.save(output_path)
    wb.close()

    logger.info("生成完了: %s (%d 件)", output_path.name, len(entries))
    return output_path


def main() -> int:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%H:%M:%S",
    )

    parser = argparse.ArgumentParser(description="xlsm → xlsx ジョブ定義抽出")
    parser.add_argument(
        "--macro-dir",
        type=Path,
        default=MACRO_DIR,
        help=f"マクロ格納フォルダ (default: {MACRO_DIR})",
    )
    args = parser.parse_args()

    try:
        generate(args.macro_dir)
    except (FileNotFoundError, RuntimeError) as e:
        logger.error("%s", e)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
