"""マクロ格納フォルダのジョブ定義を読み取る。

download_jobs.xlsx (macro_to_xlsx.py で生成) を優先し、
なければ .xlsm にフォールバックする。

Usage:
    python sf-session/macro_book_reader.py
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from .config import JOBS_XLSX, MACRO_DIR, read_ids_file

logger = logging.getLogger(__name__)

# --- 列定数（ローマ字） ---
_COL_NO = "AA"
_COL_URL = "AB"
_COL_NEW_FILENAME = "AC"
_COL_SRC_FOLDER_NAME = "AD"
_COL_ENCODE = "AE"
_COL_SKIP = "AG"

_SHEET_NAME = "SalseForce"

# データ開始行（100 = ヘッダ、101〜 = データ）
_DATA_START_ROW = 101


@dataclass
class JobEntry:
    """1行分のジョブ定義。"""

    no: str
    report_id: str | None
    has_filename: bool
    new_filename: str
    src_folder_name: str
    encode: str
    skip: str


def _col_to_index(col_letter: str) -> int:
    """Excel 列文字を 1-based インデックスに変換する。"""
    result = 0
    for ch in col_letter.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def _cell(ws, row: int, col_letter: str):
    """ワークシートからセル値を取得する。"""
    return ws.cell(row=row, column=_col_to_index(col_letter)).value


def _extract_id(url: str | None) -> str | None:
    """URL から末尾の ID を抽出する。"""
    if not url or not isinstance(url, str):
        return None
    url = url.strip()
    if "/" in url:
        return url.rsplit("/", 1)[-1]
    return url


def _has_filename(value) -> bool:
    """new_filename の指定があるか判定する。"""
    if value is None:
        return False
    s = str(value).strip()
    return bool(s) and s.isprintable()


def _find_xlsm(directory: Path) -> Path | None:
    """ディレクトリ内の .xlsm ファイルを1つ返す。"""
    files = list(directory.glob("*.xlsm"))
    if not files:
        return None
    if len(files) > 1:
        names = [f.name for f in files]
        logger.warning(".xlsm が複数あります: %s。最初のものを使用。", names)
    return files[0]


def read_jobs_from_xlsm(xlsm_path: Path) -> list[JobEntry]:
    """指定した xlsm パスからジョブ定義を読み取る。"""
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    ws = wb[_SHEET_NAME]

    entries: list[JobEntry] = []
    for row in range(_DATA_START_ROW, ws.max_row + 1):
        no = _cell(ws, row, _COL_NO)
        url = _cell(ws, row, _COL_URL)

        # No も URL も空なら終端
        if no is None and url is None:
            break

        raw_filename = _cell(ws, row, _COL_NEW_FILENAME)
        has_fn = _has_filename(raw_filename)

        entries.append(JobEntry(
            no=str(no) if no is not None else "",
            report_id=_extract_id(url),
            has_filename=has_fn,
            new_filename=str(raw_filename).strip() if has_fn else "",
            src_folder_name=str(_cell(ws, row, _COL_SRC_FOLDER_NAME) or ""),
            encode=str(_cell(ws, row, _COL_ENCODE) or ""),
            skip=str(_cell(ws, row, _COL_SKIP) or ""),
        ))

    wb.close()
    return entries


def read_jobs_from_xlsx(xlsx_path: Path) -> list[JobEntry]:
    """download_jobs.xlsx からジョブ定義を読み取る。

    Row 2: ヘッダ (no, report_id, report_name, new_filename, dst_folder_name, encode, skip)
    Row 3+: データ
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    entries: list[JobEntry] = []
    for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
        no = row[0]
        if no is None:
            break

        report_id = row[1]
        # row[2] = report_name (表示用、JobEntry では不使用)
        new_filename = row[3]
        dst_folder_name = row[4]
        encode = row[5]
        skip = row[6]

        has_fn = _has_filename(new_filename)

        entries.append(JobEntry(
            no=str(no) if no is not None else "",
            report_id=str(report_id).strip() if report_id else None,
            has_filename=has_fn,
            new_filename=str(new_filename).strip() if has_fn else "",
            src_folder_name=str(dst_folder_name or ""),
            encode=str(encode or ""),
            skip=str(skip or ""),
        ))

    wb.close()
    return entries


def read_jobs(macro_dir: Path = MACRO_DIR) -> list[JobEntry]:
    """ジョブ定義を読み取る。xlsx を優先し、なければ xlsm にフォールバック。"""
    if not macro_dir.is_dir():
        raise FileNotFoundError(f"'{macro_dir}' が見つかりません。")

    xlsx_path = macro_dir / JOBS_XLSX
    if xlsx_path.is_file():
        logger.info("xlsx から読み取り: %s", xlsx_path.name)
        return read_jobs_from_xlsx(xlsx_path)

    xlsm_path = _find_xlsm(macro_dir)
    if xlsm_path is None:
        raise FileNotFoundError(
            f"'{macro_dir.name}/' に {JOBS_XLSX} も .xlsm もありません。"
        )

    logger.info("xlsm にフォールバック: %s", xlsm_path.name)
    return read_jobs_from_xlsm(xlsm_path)


def load_active_jobs(
    macro_dir: Path = MACRO_DIR,
    *,
    ids_file: bool = False,
) -> list[JobEntry]:
    """ジョブ定義を読み込み、skip と ids-file でフィルタして返す。"""
    jobs = read_jobs(macro_dir)
    logger.info("ジョブ定義: %d 件読み取り", len(jobs))

    active = [j for j in jobs if not j.skip]
    logger.info(
        "実行対象: %d 件 (skip 除外: %d 件)",
        len(active), len(jobs) - len(active),
    )

    if ids_file:
        target_ids = read_ids_file()
        before = len(active)
        active = [j for j in active if j.report_id in target_ids]
        logger.info(
            "ids-file フィルタ: %d 件 → %d 件 (ids-file: %d IDs)",
            before, len(active), len(target_ids),
        )

    return active


def _log_jobs(entries: list[JobEntry]) -> None:
    """ジョブ定義をテーブル形式でターミナルに表示する。"""
    header = (
        f"{'No':<6} {'ID':<20} {'filename?':<10} "
        f"{'new_filename':<30} {'src_folder_name':<30} "
        f"{'encode':<10} {'skip':<6}"
    )
    lines = [header, "-" * len(header)]

    for e in entries:
        id_str = e.report_id or "(なし)"
        fn_flag = "True" if e.has_filename else "False"
        lines.append(
            f"{e.no:<6} {id_str:<20} {fn_flag:<10} "
            f"{e.new_filename:<30} {e.src_folder_name:<30} "
            f"{e.encode:<10} {e.skip:<6}"
        )

    lines.append(f"合計: {len(entries)} 件")
    logger.info("\n%s", "\n".join(lines))


if __name__ == "__main__":
    try:
        jobs = read_jobs()
    except FileNotFoundError as e:
        logger.error("%s", e)
    else:
        _log_jobs(jobs)
